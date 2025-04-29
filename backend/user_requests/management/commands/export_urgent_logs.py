# backend/user_requests/management/commands/export_urgent_logs.py

import os
import csv
from collections import defaultdict
from django.core.management.base import BaseCommand
from django.conf import settings
from user_requests.models import UrgentRequestLog
from openpyxl import Workbook
from openpyxl.styles import Font
from django.core.mail import EmailMessage


class Command(BaseCommand):
    help = 'Εξαγωγή όλων ή συγκεκριμένου μήνα/έτους urgent logs σε CSV ή Excel'

    def add_arguments(self, parser):
        parser.add_argument('--year', type=int, help='Έτος για φιλτράρισμα')
        parser.add_argument('--month', type=int, help='Μήνας για φιλτράρισμα')
        parser.add_argument('--format', type=str, choices=['csv', 'excel'], default='csv', help='Μορφή εξαγωγής: csv ή excel')

    def handle(self, *args, **options):
        year = options.get('year')
        month = options.get('month')
        export_format = options.get('format')

        logs = UrgentRequestLog.objects.all().order_by('triggered_at')

        if year and month:
            logs = logs.filter(triggered_at__year=year, triggered_at__month=month)
            self.export_logs(logs, year, month, export_format)
        else:
            grouped_logs = defaultdict(list)
            for log in logs:
                key = (log.triggered_at.year, log.triggered_at.month)
                grouped_logs[key].append(log)

            for (year, month), logs_in_month in grouped_logs.items():
                self.export_logs(logs_in_month, year, month, export_format)

    def export_logs(self, logs, year, month, export_format):
        export_dir = os.path.join(settings.BASE_DIR, 'exports', 'urgent_logs')
        os.makedirs(export_dir, exist_ok=True)

        filename_base = f"urgent_requests_{year}_{month:02d}"

        if export_format == 'csv':
            filepath = os.path.join(export_dir, filename_base + '.csv')
            with open(filepath, mode='w', newline='', encoding='utf-8') as csvfile:
                writer = csv.writer(csvfile)
                writer.writerow(['Αίτημα', 'Ημερομηνία Ενεργοποίησης', 'Υποστηρικτές'])
                for log in logs:
                    writer.writerow([
                        log.user_request.title,
                        log.triggered_at.strftime('%Y-%m-%d %H:%M'),
                        log.supporter_count
                    ])
            self.stdout.write(self.style.SUCCESS(f'✅ Δημιουργήθηκε CSV: {filepath}'))

        elif export_format == 'excel':
            filepath = os.path.join(export_dir, filename_base + '.xlsx')
            wb = Workbook()
            ws = wb.active
            ws.title = "Urgent Requests"

            # Header
            headers = ['Αίτημα', 'Ημερομηνία Ενεργοποίησης', 'Υποστηρικτές']
            ws.append(headers)

            for cell in ws[1]:
                cell.font = Font(bold=True)

            # Data rows
            for log in logs:
                ws.append([
                    log.user_request.title,
                    log.triggered_at.strftime('%Y-%m-%d %H:%M'),
                    log.supporter_count
                ])

            # Auto adjust column width
            for column_cells in ws.columns:
                max_length = max(len(str(cell.value)) for cell in column_cells if cell.value)
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column_cells[0].column_letter].width = adjusted_width

            wb.save(filepath)
            self.stdout.write(self.style.SUCCESS(f'✅ Δημιουργήθηκε Excel: {filepath}'))
        else:
            self.stdout.write(self.style.ERROR('❌ Unsupported format. Use --format csv or --format excel.'))   

        # Send email with the file attached
        self.send_email_with_attachment(filepath, year, month)
        

    def export_logs(self, logs, year, month, export_format):
        export_dir = os.path.join(settings.BASE_DIR, 'exports', 'urgent_logs')
        os.makedirs(export_dir, exist_ok=True)

        filename_base = f"urgent_requests_{year}_{month:02d}"

        if export_format == 'csv':
            filepath = os.path.join(export_dir, filename_base + '.csv')
            with open(filepath, mode='w', newline='', encoding='utf-8') as csvfile:
                writer = csv.writer(csvfile)
                writer.writerow(['Αίτημα', 'Ημερομηνία Ενεργοποίησης', 'Υποστηρικτές'])
                for log in logs:
                    writer.writerow([
                        log.user_request.title,
                        log.triggered_at.strftime('%Y-%m-%d %H:%M'),
                        log.supporter_count
                    ])
            self.stdout.write(self.style.SUCCESS(f'✅ Δημιουργήθηκε CSV: {filepath}'))

        elif export_format == 'excel':
            filepath = os.path.join(export_dir, filename_base + '.xlsx')
            wb = Workbook()
            ws = wb.active
            ws.title = "Urgent Requests"

            # Header
            headers = ['Αίτημα', 'Ημερομηνία Ενεργοποίησης', 'Υποστηρικτές']
            ws.append(headers)
            for cell in ws[1]:
                cell.font = Font(bold=True)

            # Data rows
            for log in logs:
                ws.append([
                    log.user_request.title,
                    log.triggered_at.strftime('%Y-%m-%d %H:%M'),
                    log.supporter_count
                ])

            # Auto adjust column width
            for column_cells in ws.columns:
                max_length = max(len(str(cell.value)) for cell in column_cells if cell.value)
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column_cells[0].column_letter].width = adjusted_width

            wb.save(filepath)
            self.stdout.write(self.style.SUCCESS(f'✅ Δημιουργήθηκε Excel: {filepath}'))

            # -----> Αποστολή Email με επισυναπτόμενο Excel <-----
            self.send_email_with_attachment(filepath, year, month)

    def send_email_with_attachment(self, filepath, year, month):
        subject = f"📊 Αναφορά Επειγόντων Αιτημάτων - {month:02d}/{year}"
        body = (
            f"Αγαπητέ Διαχειριστή,\n\n"
            f"Σας επισυνάπτουμε την αναφορά επειγόντων αιτημάτων για τον {month:02d}/{year}.\n\n"
            f"Με εκτίμηση,\n"
            f"Σύστημα Ψηφιακού Θυρωρού"
        )
        from_email = settings.DEFAULT_FROM_EMAIL
        to_email = ["admin@yourdomain.gr"]  # Βάλε το σωστό email ή λίστα email.

        email = EmailMessage(subject, body, from_email, to_email)
        email.attach_file(filepath)
        email.send()

        self.stdout.write(self.style.SUCCESS(f'📩 Εστάλη Email με συνημμένο: {os.path.basename(filepath)}'))
